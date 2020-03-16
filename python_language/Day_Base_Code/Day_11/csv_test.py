# -*- coding: utf-8 -*-
"""
Created on Fri Oct 18 10:45:19 2019

@author: sundooedu
"""
import csv
 
f = open('data.csv', 'r', encoding='utf-8')
rdr = csv.reader(f)
for line in rdr: #한줄씩 적출(가져온다, 읽기, 조회: 원데이터유지 추출, cf)추출:원데이터 제거 )
    print(line)
f.close()    

rdr_list=list(rdr)
print(rdr_list[0])

import csv    
f = open('output.csv', 'w', encoding='utf-8', newline='')
wr = csv.writer(f)
wr.writerow([1, "김정수", False])
wr.writerow([2, "박상미", True])
f.close()
#%%
import pandas
csv_df = pandas.read_csv("data.csv",header=None)
#%%
import numpy
a= numpy.loadtxt("data2.csv",delimiter=',')
#%%
import openpyxl

wb=openpyxl.load_workbook("score.xlsx")
ws=wb.get_sheet_by_name('시트1')
#for row in ws.rows: # 행전부
for row in ws.iter_rows(min_row=1, max_row=2): #행일부    
    row_index = row[0].row
    kor=row[0].value  # [i] i는 column 인덱스를 의미한다. 
    eng=row[1].value
    math=row[2].value
    sum= kor + eng + math #행열 인덱스를 사용하여 ws.cell(row=행인덱스, column=열인덱스)표현
    ws.cell(row=row_index, column=4).value=sum
    print(kor,eng,math,sum)
    #편집된 wb 재저장
wb.save("score.xlsx")
wb.close()    
    